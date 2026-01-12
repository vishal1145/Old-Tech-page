"""
Excel Export Utility Module
Handles exporting diagnosis results to Excel format with proper formatting.
"""
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def format_excel_worksheet(worksheet, title="Website Diagnosis Results"):
    """
    Apply professional formatting to Excel worksheet.
    
    Args:
        worksheet: openpyxl worksheet object
        title: Title for the worksheet
    """
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Format header row
    if worksheet.max_row > 0:
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border_style
    
    # Apply borders and alignment to all data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = border_style
            if cell.column_letter in ['A', 'B', 'C', 'D', 'E', 'H', 'J', 'K', 'L', 'M']:  # Text columns
                cell.alignment = wrap_alignment
            else:
                cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Set column width with some padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = max(adjusted_width, 12)


def export_single_result_to_excel(result_data, output_path=None):
    """
    Export a single diagnosis result to Excel format.
    
    Args:
        result_data: Dictionary containing diagnosis result data
        output_path: Optional path to save the Excel file. If None, generates filename.
    
    Returns:
        Path to the saved Excel file
    """
    if output_path is None:
        domain = result_data.get('domain', 'unknown')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_domain = domain.replace('.', '_').replace('/', '_')[:30]
        filename = f"diagnosis_{safe_domain}_{timestamp}.xlsx"
        output_path = os.path.join('results', filename)
    
    # Ensure results directory exists
    os.makedirs('results', exist_ok=True)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Main Overview Sheet
        overview_data = {
            'Field': [
                'URL',
                'Domain',
                'Technology',
                'Status',
                'Load Time',
                'First Contentful Paint (ms)',
                'Console Error Count',
                'Vulnerability Detected',
                'Vulnerabilities Count',
                'Diagnosis Date'
            ],
            'Value': [
                result_data.get('url', 'N/A'),
                result_data.get('domain', 'N/A'),
                result_data.get('tech', 'Unknown'),
                result_data.get('status', 'unknown'),
                result_data.get('load_time', 'N/A'),
                result_data.get('first_contentful_paint_ms', 'N/A'),
                result_data.get('console_error_count', 0),
                'Yes' if result_data.get('vulnerability_detected', False) else 'No',
                len(result_data.get('vulnerabilities', [])),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
        }
        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, sheet_name='Overview', index=False)
        
        # Technical Observation Sheet
        if result_data.get('technical_observation'):
            observation_data = {
                'Technical Observation': [result_data.get('technical_observation')]
            }
            observation_df = pd.DataFrame(observation_data)
            observation_df.to_excel(writer, sheet_name='Technical Observation', index=False)
        
        # Vulnerabilities Sheet
        vulnerabilities = result_data.get('vulnerabilities', [])
        if vulnerabilities:
            vuln_data = []
            for vuln in vulnerabilities:
                vuln_data.append({
                    'Type': vuln.get('type', 'N/A'),
                    'Version': vuln.get('version', 'unknown'),
                    'Matched Text': vuln.get('matched_text', '')[:200]  # Limit length
                })
            vuln_df = pd.DataFrame(vuln_data)
            vuln_df.to_excel(writer, sheet_name='Vulnerabilities', index=False)
        else:
            # Create empty sheet with message
            empty_df = pd.DataFrame({'Message': ['No vulnerabilities detected']})
            empty_df.to_excel(writer, sheet_name='Vulnerabilities', index=False)
        
        # Console Errors Sheet
        console_errors = result_data.get('console_errors', [])
        if console_errors:
            error_data = {
                'Error Number': range(1, len(console_errors) + 1),
                'Error Message': console_errors
            }
            error_df = pd.DataFrame(error_data)
            error_df.to_excel(writer, sheet_name='Console Errors', index=False)
        else:
            # Create empty sheet with message
            empty_df = pd.DataFrame({'Message': ['No console errors detected']})
            empty_df.to_excel(writer, sheet_name='Console Errors', index=False)
    
    # Apply formatting to all sheets
    workbook = load_workbook(output_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        format_excel_worksheet(worksheet, title=sheet_name)
    
    workbook.save(output_path)
    return output_path


def export_bulk_results_to_excel(results_list, output_path=None):
    """
    Export multiple diagnosis results to a single Excel file with summary sheet.
    
    Args:
        results_list: List of diagnosis result dictionaries
        output_path: Optional path to save the Excel file. If None, generates filename.
    
    Returns:
        Path to the saved Excel file
    """
    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"bulk_diagnosis_results_{timestamp}.xlsx"
        output_path = os.path.join('results', filename)
    
    # Ensure results directory exists
    os.makedirs('results', exist_ok=True)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Summary Sheet
        summary_data = []
        for idx, result in enumerate(results_list, 1):
            summary_data.append({
                'No.': idx,
                'URL': result.get('url', 'N/A'),
                'Domain': result.get('domain', 'N/A'),
                'Technology': result.get('tech', 'Unknown'),
                'Status': result.get('status', 'unknown'),
                'Load Time': result.get('load_time', 'N/A'),
                'FCP (ms)': result.get('first_contentful_paint_ms', 'N/A'),
                'Console Errors': result.get('console_error_count', 0),
                'Vulnerabilities': len(result.get('vulnerabilities', [])),
                'Vulnerability Detected': 'Yes' if result.get('vulnerability_detected', False) else 'No',
                'Technical Observation': result.get('technical_observation', 'N/A')
            })
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Individual result sheets
        for idx, result in enumerate(results_list, 1):
            domain = result.get('domain', f'result_{idx}')
            safe_domain = domain.replace('.', '_').replace('/', '_')[:25]
            sheet_name = f"{idx}_{safe_domain}"[:31]  # Excel sheet name limit
            
            # Overview for this result
            overview_data = {
                'Field': [
                    'URL',
                    'Domain',
                    'Technology',
                    'Status',
                    'Load Time',
                    'First Contentful Paint (ms)',
                    'Console Error Count',
                    'Vulnerability Detected',
                    'Vulnerabilities Count'
                ],
                'Value': [
                    result.get('url', 'N/A'),
                    result.get('domain', 'N/A'),
                    result.get('tech', 'Unknown'),
                    result.get('status', 'unknown'),
                    result.get('load_time', 'N/A'),
                    result.get('first_contentful_paint_ms', 'N/A'),
                    result.get('console_error_count', 0),
                    'Yes' if result.get('vulnerability_detected', False) else 'No',
                    len(result.get('vulnerabilities', []))
                ]
            }
            overview_df = pd.DataFrame(overview_data)
            overview_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Apply formatting to all sheets
    workbook = load_workbook(output_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        format_excel_worksheet(worksheet, title=sheet_name)
    
    workbook.save(output_path)
    return output_path


def export_company_list_to_excel(results_list, output_path=None):
    """
    Export all company diagnosis results to a single Excel file with complete details.
    Creates a comprehensive single-sheet export with all diagnosis data.
    
    Args:
        results_list: List of diagnosis result dictionaries
        output_path: Optional path to save the Excel file. If None, generates filename.
    
    Returns:
        Path to the saved Excel file
    """
    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"company_diagnosis_export_{timestamp}.xlsx"
        output_path = os.path.join('results', filename)
    
    # Ensure results directory exists
    os.makedirs('results', exist_ok=True)
    
    # Prepare data for Excel
    export_data = []
    
    for result in results_list:
        # Format vulnerabilities as comma-separated list
        vulnerabilities = result.get('vulnerabilities', [])
        vuln_list = ', '.join([f"{v.get('type', 'N/A')} (v{v.get('version', 'unknown')})" 
                              for v in vulnerabilities]) if vulnerabilities else 'None'
        
        # Format console errors (truncated summary)
        console_errors = result.get('console_errors', [])
        if console_errors:
            # Show first 3 errors, truncate if too long
            error_summary = []
            for i, error in enumerate(console_errors[:3]):
                truncated_error = error[:100] + '...' if len(error) > 100 else error
                error_summary.append(f"{i+1}. {truncated_error}")
            if len(console_errors) > 3:
                error_summary.append(f"... and {len(console_errors) - 3} more errors")
            console_errors_text = ' | '.join(error_summary)
        else:
            console_errors_text = 'None'
        
        # Get diagnosis date from file modification time or use current date
        diagnosis_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if 'modified' in result:
            try:
                diagnosis_date = datetime.fromtimestamp(result['modified']).strftime('%Y-%m-%d %H:%M:%S')
            except:
                pass
        
        export_data.append({
            'Domain': result.get('domain', 'N/A'),
            'URL': result.get('url', 'N/A'),
            'Technology': result.get('tech', 'Unknown'),
            'Status': result.get('status', 'unknown'),
            'Load Time': result.get('load_time', 'N/A'),
            'FCP (ms)': result.get('first_contentful_paint_ms', 'N/A'),
            'Console Errors Count': result.get('console_error_count', 0),
            'Console Errors': console_errors_text,
            'Vulnerabilities Count': len(vulnerabilities),
            'Vulnerabilities': vuln_list,
            'Vulnerability Detected': 'Yes' if result.get('vulnerability_detected', False) else 'No',
            'Technical Observation': result.get('technical_observation', 'N/A'),
            'Diagnosis Date': diagnosis_date
        })
    
    # Create DataFrame
    df = pd.DataFrame(export_data)
    
    # Create Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Company Diagnosis List', index=False)
    
    # Apply formatting
    workbook = load_workbook(output_path)
    worksheet = workbook['Company Diagnosis List']
    format_excel_worksheet(worksheet, title='Company Diagnosis List')
    
    # Set specific column widths for better readability
    column_widths = {
        'A': 25,  # Domain
        'B': 40,  # URL
        'C': 20,  # Technology
        'D': 12,  # Status
        'E': 12,  # Load Time
        'F': 12,  # FCP (ms)
        'G': 18,  # Console Errors Count
        'H': 50,  # Console Errors
        'I': 20,  # Vulnerabilities Count
        'J': 40,  # Vulnerabilities
        'K': 20,  # Vulnerability Detected
        'L': 60,  # Technical Observation
        'M': 20   # Diagnosis Date
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    workbook.save(output_path)
    return output_path
